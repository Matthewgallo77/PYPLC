import openpyxl
import struct
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import snap7 # IMPORT SNAP7 LIBRARY
import re
import xlwings as xw
from snap7.util import *
from sys import getsizeof # IMPORT TO GET SIZE OF DATA
from snap7.common import check_error, ipv4, load_library
from snap7.types import S7WLBit, S7WLByte, S7WLInt, S7WLReal, S7WLTimer
from win32com.client import Dispatch # pip install pywin32

# DOCUMENTATION: https://python-snap7.readthedocs.io/en/1.0/client.html#snap7.client.Client.as_download 

# CREATE A COPY OF TAG TABLE WITH DESIRED TAGS

offsets = [['Bool',4],['Int',2],['Real',4]]

# while True:
#         try:
#             path = input("Enter the Path of the excel file: ")
#             wb = xw.Book(path, read_only=True)
#             return [wb, path]
#         except snap7.snap7exceptions:
#             print("Path was entered incorrectly. Please check if xlsx file is in same folder as EXE file or ENTER AGAIN: ")

def connectPLC():

    while True:
        try:
            IP = input("Enter the IP address of the PLC: ")
            PLC.connect(IP,0,1) # IP, RACK #, SLOT #
            print("CONNECTION STATUS: \nPLC: True") # DISPLAYS IF CONNECTION TO PLC IS VALID
            return IP
        except snap7.error:
            print("CONNECTION STATUS: \nPLC: False\nCheck PLC connection or re-enter IP.")

def adjustXL():
    # ORGANIZES EXCEL SHEET FOR READABILITY
    
    sht = wb.sheets[0]
    if sht.range('G1').value is not None:
        column_range = sht.range('E:J')
        column_range.api.Delete()
        sht.range('E1').value = 'Values'
        sht.range('F1').value = 'PLC Status: ' + str(PLC.get_connected())
        with xw.App(visible=False) as app:
            for ws in wb.sheets:
                ws.autofit(axis='columns')

            sht.range('E1').column_width = 15

def getTagInfo():

    lastRow = str(len(list(uiSheet.rows))) # GETS THE LAST NON EMPTY ROW
    NAMES, DATATYPES, ADDRESSES = uiSheet['A2':'A'+lastRow], uiSheet['C2':'C'+lastRow], uiSheet['D2':'D'+lastRow] # ACCESS NECESSARY COLUMNS
    nameList, datatypeList, addressList = [], [], [] # INITIALIZE LISTS
    nameList, datatypeList, addressList  = dataList(NAMES, nameList), dataList(DATATYPES, datatypeList), dataList(ADDRESSES, addressList)
    datatypeList = [offsets[0] if dataType == 'Bool' else offsets[1] if dataType == 'Int' else offsets[2] for dataType in datatypeList]
    address_dataType = {addressList[i]: datatypeList[i] for i in range(len(addressList))} # CONTAINS {ADDRESS: DATATYPE}

    return address_dataType

def dataList(columnInput, infoList):
        infoList = []
        for CELL in columnInput:
            for INFO in CELL:
                infoList.append(INFO.value)
        return infoList

def ReadTags(): # READS ALL TAGS AND RETRIEVES THERE VALUE
    valueList = []
    addressList = []

    for key, value in address_dataType.items(): # ADDRESS : DATATYPE
        if 'I' in key: # MERKER 
            valueList.append(readInput(key, value)) # PROCESS INPUTS - OFFSET, DATATYPE
        elif 'M' in key:
            valueList.append(readMerker(key,value)) # PROCESS MERKERS - OFFSET, DATATYPE
        elif 'Q' in key:
            valueList.append(readOutput(key,value)) # PROCESS OUTPUT - OFFSET, DATATYPE
        else:
            print("Data type not anticipated")

    return valueList

def readInput(offset, data_type):
    offset = re.sub("[^\d\.]", "", offset).split('.')
    if len(offset) == 1:
        offset.append(0)
    offset = [int(index) for index in offset]
    byteArray = PLC.read_area(snap7.types.Areas.PE, 0, offset[0], data_type[1])
    return ReadData(offset[0], offset[1], data_type, byteArray)

def readMerker(offset, data_type):
    offset = re.sub("[^\d\.]", "", offset).split('.')
    if len(offset) == 1:
        offset.append(0)
    offset = [int(index) for index in offset]
    byteArray = PLC.read_area(snap7.types.Areas.MK, 0, offset[0], data_type[1])
    return ReadData(offset[0], offset[1], data_type, byteArray)

def readOutput(offset, data_type):
    offset = re.sub("[^\d\.]", "", offset).split('.')
    if len(offset) == 1:
        offset.append(0)
    offset = [int(index) for index in offset]
    byteArray = PLC.read_area(snap7.types.Areas.PA, 0, offset[0], data_type[1])
    return ReadData(offset[0], offset[1], data_type, byteArray)

def ReadData(byte, bit, datatype, byteArray): # M, MB, MW, MD
    if datatype[0] == 'Real':
        return get_real(byteArray, 0)
    # elif datatype[0] == 'Bool':
        # return get_bool(byteArray, byte, bit)
    elif datatype[0] == 'Int':
        return get_int(byteArray, 0)
    else:
        return 'ERROR: Data type has not been anticipated'

def write_toExcel():
    
    sheet = xw.sheets[0]
  
    rowCount=2
    lastRow = len(list(uiSheet.rows))
    while (PLC.get_connected() and PLC.get_cpu_state() == 'S7CpuStatusRun'):
        valueList = ReadTags()
        rowCount = 2 # RESET
        while rowCount<=lastRow: 
            for value in valueList:
                sheet.range('E'+str(rowCount)).value = value
                rowCount+=1
    
    if (PLC.get_cpu_state() != 'S7CpuStatusRun'):
        sheet.range('F1').value = 'PLC Connection: False'

def setPath():
    while True:
        try:
            path = input("Enter the Path of the excel file: ")
            wb = xw.Book(path, read_only=True)
            return [wb, path]
        except FileNotFoundError:
            print("Path was entered incorrectly. Please check if xlsx file is in same folder as EXE file or ENTER AGAIN: ")
      
    
if __name__ == "__main__":
    PLC = snap7.client.Client()
    IP = connectPLC()
    excelSheetData = setPath()
    wb, path = excelSheetData[0], excelSheetData[1]
    excelWorkbook = load_workbook(path)
    adjustXL() # ADJUSTS XL FOR READABILITY
    uiSheet = excelWorkbook.active 
    address_dataType = getTagInfo()
    valueList = ReadTags()
    write_toExcel()
