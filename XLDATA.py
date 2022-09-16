


import openpyxl
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import snap7 # IMPORT SNAP7 LIBRARY
import re
import xlwings as xw
from snap7.util import *
from sys import getsizeof # IMPORT TO GET SIZE OF DATA
from snap7.common import check_error, ipv4, load_library
from snap7.types import S7WLBit, S7WLByte, S7WLInt, S7WLReal, S7WLTimer
from win32com.client import Dispatch # pip install pywin32

# DOCUMENTATION: https://python-snap7.readthedocs.io/en/1.0/client.html#snap7.client.Client.as_download 

def getTagInfo():
    
    lastRow = str(len(list(uiSheet.rows))) # GETS THE LAST NON EMPTY ROW
    NAMES, DATATYPES, ADDRESSES = uiSheet['A2':'A'+lastRow], uiSheet['C2':'C'+lastRow], uiSheet['D2':'D'+lastRow] # ACCESS NECESSARY COLUMNS
    nameList, datatypeList, addressList = [], [], [] # INITIALIZE LISTS

    nameList, datatypeList, addressList  = dataList(NAMES, nameList), dataList(DATATYPES, datatypeList), dataList(ADDRESSES, addressList)
    datatypeList = [snap7.types.S7WLBit if dataType == 'Bool' else snap7.types.S7WLWord if dataType == 'Word' else snap7.types.S7WLInt if dataType == 'Int' else snap7.types.S7WLReal if dataType == 'Real' else snap7.types.S7WLTimer for dataType in datatypeList]
    address_dataType = {addressList[i]: datatypeList[i] for i in range(len(addressList))} # CONTAINS {ADDRESS: DATATYPE}
    
    return address_dataType

def dataList(columnInput, infoList):
        infoList = []
        for CELL in columnInput:
            for INFO in CELL:
                infoList.append(INFO.value)
        return infoList

def ReadData(PLC, byte, bit, datatype, byteArray): # M, MB, MW, MD
    if datatype == snap7.types.S7WLBit:
        return get_bool(byteArray, byte, bit)
    elif datatype == snap7.types.S7WLInt or datatype == snap7.types.S7WLWord:
        return get_int(byteArray, byte)
    elif datatype == snap7.types.S7WLReal:
        return get_real(byteArray, byte)
    elif datatype == snap7.types.S7WLTimer:
        return get_time(byteArray, byte)
    else:
        return 'ERROR: Data type has not been anticipated'


def ReadTags(PLC): # READS ALL TAGS AND RETRIEVES THERE VALUE
    valueList = []
    addressList = []

    for key, value in address_dataType.items(): # ADDRESS : DATATYPE
        if 'M' in key: # MERKER  
            valueList.append(readMerker(key, value))
        elif 'I' in key: # INPUT
            valueList.append(readInput(key, value))
        elif 'Q' in key: # OUTPUT
            valueList.append(readOutput(key, value))

    return valueList

def readMerker(offset, data_type):
    offset = re.sub("[^\d\.]", "", offset).split('.')
    if len(offset) == 1:
        offset.append(0)
    offset = [int(index) for index in offset]
    byte_array = PLC1.mb_read(offset[0], data_type)

    return ReadData(PLC1, offset[0], offset[1], data_type , byte_array)
    
def readInput(offset, data_type):
    offset = re.sub("[^\d\.]", "", offset).split('.')
    
    if len(offset) == 1:
        offset.append(0)
    offset = [int(index) for index in offset]
    byte_array = PLC1.eb_read(offset[0], data_type)

    return ReadData(PLC1, offset[0], offset[1], data_type, byte_array)

def readOutput(offset, data_type):
    offset = re.sub("[^\d\.]", "", offset).split('.')
    if len(offset) == 1:
        offset.append(0)
    offset = [int(index) for index in offset]
    byte_array = PLC1.ab_read(offset[0], data_type)
    
    return ReadData(PLC1, offset[0], offset[1], data_type, byte_array)

def write_toExcel(PLC):
    
    wb = xw.Book('PLCTags.xlsx')
    sheet = xw.sheets[0]
    rowCount=2
    lastRow = len(list(uiSheet.rows))

    while PLC.get_connected():
        rowCount = 2 # RESET
        while rowCount<=lastRow: 
            valueList = ReadTags(PLC1)
            for value in valueList:
                sheet.range('K'+str(rowCount)).value = value
                rowCount+=1

def adjustXL():

    wb = xw.Book(path)
    # DELETE UNECCESSARY COLUMNS
    with xw.App(visible=False) as app:
        for ws in wb.sheets:
            ws.autofit(axis='columns')

        wb.save(path)

    

def connectionStatus(PLC):
    try:
        PLC1.connect('192.168.0.1',0,1) # IP, RACK #, SLOT #
        print("CONNECTION STATUS: \n" + "PLC1: " + str(PLC1.get_connected())) # DISPLAYS IF CONNECTION TO PLC IS VALID
    except:
        print("CONNECTION STATUS: \n" "PLC1: " + str(PLC1.get_connected()))

if __name__ == "__main__":

    path = 'PLCTags.xlsx' # FILE NAME
    excelWorkbook = load_workbook(path)
    adjustXL() # ADJUSTS FORMATTING OF EXCEL SHEET
    uiSheet = excelWorkbook.active
    address_dataType = getTagInfo()
    PLC1 = snap7.client.Client()
    
    x=ReadTags(PLC1)
    print(x)

